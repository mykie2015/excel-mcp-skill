#!/usr/bin/env python3
"""Apply a formatting template to an Excel workbook using openpyxl."""

from __future__ import annotations

import argparse
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException


PRESETS = {
    "report": {
        "header_fill": "1F4E78",
        "header_font": "FFFFFF",
        "border": "B8CCE4",
        "int_fmt": "#,##0",
        "float_fmt": "#,##0.00",
        "date_fmt": "yyyy-mm-dd",
        "percent_fmt": "0.00%",
    },
    "dashboard": {
        "header_fill": "2E7D32",
        "header_font": "FFFFFF",
        "border": "A5D6A7",
        "int_fmt": "#,##0",
        "float_fmt": "#,##0.0",
        "date_fmt": "mmm d, yyyy",
        "percent_fmt": "0.0%",
    },
    "data-table": {
        "header_fill": "455A64",
        "header_font": "FFFFFF",
        "border": "CFD8DC",
        "int_fmt": "#,##0",
        "float_fmt": "#,##0.00",
        "date_fmt": "yyyy-mm-dd",
        "percent_fmt": "0.00%",
    },
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Apply formatting template to an XLSX workbook: styled headers, "
            "auto column widths, frozen header row, and inferred number formats."
        )
    )
    parser.add_argument("input_xlsx", help="Path to input .xlsx file")
    parser.add_argument("output_xlsx", help="Path to output .xlsx file")
    parser.add_argument(
        "--preset",
        default="report",
        choices=sorted(PRESETS.keys()),
        help="Style preset to apply (default: report)",
    )
    return parser.parse_args()


def is_percent_header(header: Any) -> bool:
    if header is None:
        return False
    text = str(header).strip().lower()
    return any(token in text for token in ("%", "percent", "pct", "rate"))


def infer_column_number_format(values: list[Any], header: Any, preset: dict[str, str]) -> str | None:
    if is_percent_header(header):
        return preset["percent_fmt"]

    for value in values:
        if value is None:
            continue
        if isinstance(value, bool):
            return None
        if isinstance(value, int):
            return preset["int_fmt"]
        if isinstance(value, float):
            return preset["float_fmt"]
        if isinstance(value, (datetime, date)):
            return preset["date_fmt"]
        return None

    return None


def apply_header_style(cell: Cell, preset: dict[str, str], border: Border) -> None:
    cell.font = Font(bold=True, color=preset["header_font"])
    cell.fill = PatternFill(fill_type="solid", fgColor=preset["header_fill"])
    cell.border = border


def autosize_columns(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            value = row[0].value
            if value is None:
                continue
            length = len(str(value))
            if length > max_length:
                max_length = length
        adjusted = min(max(max_length + 2, 10), 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted


def apply_sheet_formatting(ws, preset: dict[str, str]) -> None:
    if ws.max_row < 1 or ws.max_column < 1:
        return

    thin_side = Side(style="thin", color=preset["border"])
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col_idx in range(1, ws.max_column + 1):
        header_cell = ws.cell(row=1, column=col_idx)
        if header_cell.value is not None:
            apply_header_style(header_cell, preset, border)

        data_values = [
            ws.cell(row=row_idx, column=col_idx).value
            for row_idx in range(2, ws.max_row + 1)
        ]
        number_fmt = infer_column_number_format(data_values, header_cell.value, preset)
        if number_fmt:
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    cell.number_format = number_fmt

    ws.freeze_panes = "A2"
    autosize_columns(ws)


def main() -> int:
    args = parse_args()
    input_path = Path(args.input_xlsx)
    output_path = Path(args.output_xlsx)
    preset = PRESETS[args.preset]

    if not input_path.exists():
        print(f"Error: input workbook not found: {input_path}", file=sys.stderr)
        return 1

    if input_path.suffix.lower() != ".xlsx":
        print("Error: input workbook must be a .xlsx file", file=sys.stderr)
        return 1

    try:
        workbook = load_workbook(filename=str(input_path))
    except FileNotFoundError:
        print(f"Error: input workbook not found: {input_path}", file=sys.stderr)
        return 1
    except (InvalidFileException, BadZipFile, OSError) as exc:
        print(f"Error: could not open workbook '{input_path}': {exc}", file=sys.stderr)
        return 1

    if not workbook.worksheets:
        print("Warning: workbook has no worksheets. Saving output unchanged.", file=sys.stderr)
    else:
        for ws in workbook.worksheets:
            if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
                continue
            apply_sheet_formatting(ws, preset)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        workbook.save(str(output_path))
    except OSError as exc:
        print(f"Error: failed to save workbook to '{output_path}': {exc}", file=sys.stderr)
        return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
