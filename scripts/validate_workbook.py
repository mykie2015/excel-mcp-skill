#!/usr/bin/env python3
"""Validate an .xlsx workbook and emit a JSON summary."""

from __future__ import annotations

import argparse
import json
import sys
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException


@dataclass(frozen=True)
class TypeProfile:
    expected: str | None
    mismatches: int
    sampled_non_empty: int


def cell_type(value: Any) -> str:
    if value is None:
        return "empty"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, int):
        return "int"
    if isinstance(value, float):
        return "float"
    if isinstance(value, str):
        return "str"
    return type(value).__name__


def normalize_for_dupe(value: Any) -> Any:
    if isinstance(value, str):
        return value.strip()
    return value


def infer_expected_type(values: list[Any]) -> TypeProfile:
    non_empty_types = [cell_type(v) for v in values if v is not None and not (isinstance(v, str) and v.strip() == "")]
    if not non_empty_types:
        return TypeProfile(expected=None, mismatches=0, sampled_non_empty=0)

    expected, _ = Counter(non_empty_types).most_common(1)[0]
    mismatches = sum(1 for t in non_empty_types if t != expected)
    return TypeProfile(expected=expected, mismatches=mismatches, sampled_non_empty=len(non_empty_types))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Validate an .xlsx workbook. Reports sheets, headers, row/empty/duplicate counts, "
            "and type mismatches as JSON."
        )
    )
    parser.add_argument("workbook", help="Path to the .xlsx workbook file")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    workbook_path = Path(args.workbook)

    summary: dict[str, Any] = {
        "workbook_path": str(workbook_path),
        "valid": False,
        "errors": [],
        "sheets": [],
    }

    if not workbook_path.exists():
        summary["errors"].append(f"File not found: {workbook_path}")
        print(json.dumps(summary, indent=2))
        return 1

    if workbook_path.suffix.lower() != ".xlsx":
        summary["errors"].append("Only .xlsx files are supported.")
        print(json.dumps(summary, indent=2))
        return 1

    try:
        wb = load_workbook(filename=workbook_path, data_only=False, read_only=True)
    except FileNotFoundError:
        summary["errors"].append(f"File not found: {workbook_path}")
        print(json.dumps(summary, indent=2))
        return 1
    except (InvalidFileException, BadZipFile, OSError, ValueError) as exc:
        summary["errors"].append(f"Failed to open workbook: {exc}")
        print(json.dumps(summary, indent=2))
        return 1

    validation_errors: list[str] = []

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        headers_row = rows[0] if rows else ()
        max_columns = max((len(r) for r in rows), default=0)
        data_rows = rows[1:] if len(rows) > 1 else []

        headers: list[str] = []
        for col_idx in range(max_columns):
            raw = headers_row[col_idx] if col_idx < len(headers_row) else None
            if raw is None or (isinstance(raw, str) and raw.strip() == ""):
                headers.append(f"Column_{get_column_letter(col_idx + 1)}")
            else:
                headers.append(str(raw))

        non_empty_data_rows = [
            row
            for row in data_rows
            if any(cell is not None and not (isinstance(cell, str) and cell.strip() == "") for cell in row)
        ]

        empty_cells = 0
        duplicate_rows = 0
        seen_rows: set[tuple[Any, ...]] = set()
        for row in non_empty_data_rows:
            normalized = tuple(normalize_for_dupe(v) for v in row)
            if normalized in seen_rows:
                duplicate_rows += 1
            else:
                seen_rows.add(normalized)

            for cell in row:
                if cell is None or (isinstance(cell, str) and cell.strip() == ""):
                    empty_cells += 1

        type_mismatches_total = 0
        columns: list[dict[str, Any]] = []
        for col_idx in range(max_columns):
            col_values = [row[col_idx] if col_idx < len(row) else None for row in non_empty_data_rows]
            profile = infer_expected_type(col_values)
            type_mismatches_total += profile.mismatches
            columns.append(
                {
                    "header": headers[col_idx],
                    "expected_type": profile.expected,
                    "mismatch_count": profile.mismatches,
                    "sampled_non_empty": profile.sampled_non_empty,
                }
            )

        if duplicate_rows > 0:
            validation_errors.append(f"Sheet '{ws.title}' has {duplicate_rows} duplicate data row(s).")
        if type_mismatches_total > 0:
            validation_errors.append(f"Sheet '{ws.title}' has {type_mismatches_total} type mismatch(es).")

        summary["sheets"].append(
            {
                "name": ws.title,
                "headers": headers,
                "row_count": len(non_empty_data_rows),
                "empty_cells": empty_cells,
                "duplicate_rows": duplicate_rows,
                "type_mismatches": type_mismatches_total,
                "columns": columns,
            }
        )

    try:
        wb.close()
    except Exception:
        pass

    if validation_errors:
        summary["errors"].extend(validation_errors)
        summary["valid"] = False
        print(json.dumps(summary, indent=2))
        return 1

    summary["valid"] = True
    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
