#!/usr/bin/env python3
"""Create a pivot table sheet inside an existing .xlsx workbook."""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

ALLOWED_AGGREGATIONS = {
    "sum": "sum",
    "count": "count",
    "average": "mean",
    "min": "min",
    "max": "max",
}


def parse_csv_fields(value: str) -> list[str]:
    fields = [item.strip() for item in value.split(",") if item.strip()]
    if not fields:
        raise argparse.ArgumentTypeError("At least one field is required.")
    return fields


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Generate a pivot table from an existing worksheet and write the result "
            "to a new sheet in the same workbook."
        )
    )
    parser.add_argument("workbook", help="Path to an input .xlsx workbook")
    parser.add_argument("--sheet", required=True, help="Source worksheet name")
    parser.add_argument(
        "--rows",
        required=True,
        type=parse_csv_fields,
        help="Comma-separated field names for pivot rows (index)",
    )
    parser.add_argument(
        "--columns",
        required=True,
        type=parse_csv_fields,
        help="Comma-separated field names for pivot columns",
    )
    parser.add_argument(
        "--values",
        required=True,
        type=parse_csv_fields,
        help="Comma-separated numeric/value field names to aggregate",
    )
    parser.add_argument(
        "--agg",
        required=True,
        choices=sorted(ALLOWED_AGGREGATIONS.keys()),
        help="Aggregation type: sum|count|average|min|max",
    )
    parser.add_argument(
        "--output-sheet",
        default="PivotTable",
        help="Name for the output worksheet (default: PivotTable)",
    )
    return parser.parse_args()


def validate_args(args: argparse.Namespace) -> None:
    workbook_path = Path(args.workbook)
    if not workbook_path.exists():
        raise ValueError(f"File not found: {workbook_path}")
    if workbook_path.suffix.lower() != ".xlsx":
        raise ValueError("Only .xlsx files are supported.")

    if args.output_sheet == args.sheet:
        raise ValueError("Output sheet name must differ from source sheet name.")

    duplicate_groups = {
        "rows": args.rows,
        "columns": args.columns,
        "values": args.values,
    }
    for group_name, fields in duplicate_groups.items():
        if len(fields) != len(set(fields)):
            raise ValueError(f"Duplicate fields are not allowed in --{group_name}.")


def load_source_dataframe(workbook_path: Path, sheet_name: str) -> pd.DataFrame:
    try:
        return pd.read_excel(workbook_path, sheet_name=sheet_name, engine="openpyxl")
    except ValueError as exc:
        raise ValueError(f"Failed to read sheet '{sheet_name}': {exc}") from exc


def validate_fields(df: pd.DataFrame, args: argparse.Namespace) -> None:
    required = set(args.rows + args.columns + args.values)
    missing = sorted(required - set(df.columns))
    if missing:
        raise ValueError(f"Missing field(s) in source sheet: {', '.join(missing)}")


def build_pivot(df: pd.DataFrame, args: argparse.Namespace) -> pd.DataFrame:
    aggfunc = ALLOWED_AGGREGATIONS[args.agg]
    pivot_df = pd.pivot_table(
        df,
        index=args.rows,
        columns=args.columns,
        values=args.values,
        aggfunc=aggfunc,
        dropna=False,
    )

    if isinstance(pivot_df.columns, pd.MultiIndex):
        pivot_df.columns = [
            " | ".join(str(item) for item in col if item is not None and str(item) != "")
            for col in pivot_df.columns.to_list()
        ]

    return pivot_df.reset_index()


def write_pivot_sheet(workbook_path: Path, output_sheet: str, pivot_df: pd.DataFrame) -> None:
    try:
        wb = load_workbook(workbook_path)
    except (InvalidFileException, OSError, ValueError) as exc:
        raise ValueError(f"Failed to open workbook for writing: {exc}") from exc

    if output_sheet in wb.sheetnames:
        del wb[output_sheet]

    ws = wb.create_sheet(title=output_sheet)

    for col_idx, header in enumerate(pivot_df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=str(header))

    for row_idx, row_values in enumerate(pivot_df.itertuples(index=False, name=None), start=2):
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(workbook_path)


def main() -> int:
    args = parse_args()

    try:
        validate_args(args)
        workbook_path = Path(args.workbook)
        df = load_source_dataframe(workbook_path, args.sheet)
        validate_fields(df, args)
        pivot_df = build_pivot(df, args)
        write_pivot_sheet(workbook_path, args.output_sheet, pivot_df)
    except ValueError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1
    except Exception as exc:  # Guardrail for unexpected runtime failures.
        print(f"Unexpected error: {exc}", file=sys.stderr)
        return 1

    print(
        f"Pivot table written to sheet '{args.output_sheet}' in '{args.workbook}' "
        f"using aggregation '{args.agg}'."
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
